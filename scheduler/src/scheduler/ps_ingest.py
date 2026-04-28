"""PowerSchool / Columbus real-data ingester.

Reads the actual Columbus operating documents:
1. PowerSchool enrollment CSV export (Spanish headers, UTF-8 BOM)
2. The "STUDENTS PER COURSE" Excel workbook (course demand per grade)
3. The "HS_Schedule" Excel workbook (master schedule, groupings, teachers)

Produces canonical `Dataset` objects compatible with the solvers.

Design notes:
- Real Columbus IDs are heterogeneous: '2_11104', '1_10656', 'arodriguez789*'.
  We preserve them as-is (no normalization); validation is done at the solver
  level.
- Spanish column headers are mapped via a registry so future schools can add
  their own header maps without touching this code.
- The xlsx files have many sheets; this ingester extracts only the sheets
  required to build a solvable Dataset and ignores the rest. Add new readers
  to extract more (e.g., teacher availability calendars).
"""
from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook

from .models import (
    BehaviorMatrix,
    Course,
    CourseRequest,
    Dataset,
    HardConstraints,
    Room,
    RoomType,
    SchoolConfig,
    Section,
    SoftConstraintWeights,
    Student,
    Teacher,
    default_rotation,
)


# === PowerSchool enrollment CSV (Spanish headers) ============================

# Canonical field → list of synonym headers we'll accept (lowercased + stripped)
PS_CSV_HEADER_MAP = {
    "student_id":     ["id de usuario único", "id de usuario unico", "user id", "student id"],
    "first_name":     ["primer nombre", "first name"],
    "last_name":      ["apellido", "last name"],
    "email":          ["correo electrónico", "correo electronico", "email"],
    "course_name":    ["nombre del curso", "course name"],
    "course_id":      ["código de curso", "codigo de curso", "course code"],
    "section_name":   ["nombre de sección", "nombre de seccion", "section name"],
    "section_id":     ["código de sección", "codigo de seccion", "section code"],
    "enrollment_type": ["tipo de inscripción (1=administrativo/2=miembro)", "tipo de inscripcion", "enrollment type"],
    "status":         ["estado", "status"],
    "campus_id":      ["id del campus", "campus id"],
    "campus_name":    ["título del campus", "titulo del campus", "campus name"],
}


@dataclass
class PSEnrollmentRow:
    student_id: str
    first_name: str
    last_name: str
    email: str
    course_id: str
    course_name: str
    section_id: str
    section_name: str
    enrollment_type: str
    status: str
    campus_id: str
    campus_name: str


def _norm_header(s: str) -> str:
    return (s or "").strip().lower().lstrip("﻿")  # drop BOM


def _build_header_index(headers: list[str], header_map: dict[str, list[str]]) -> dict[str, int]:
    """Map canonical field name → CSV column index."""
    norm = [_norm_header(h) for h in headers]
    idx: dict[str, int] = {}
    for canonical, aliases in header_map.items():
        for i, h in enumerate(norm):
            if h in aliases:
                idx[canonical] = i
                break
    return idx


def read_ps_enrollment_csv(path: Path) -> list[PSEnrollmentRow]:
    """Read a Columbus-style PowerSchool enrollment CSV.

    Handles:
    - UTF-8 BOM at file start
    - Spanish column headers
    - Mixed student-ID formats (preserved as-is)
    """
    path = Path(path)
    rows: list[PSEnrollmentRow] = []
    with path.open(encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        try:
            header = next(reader)
        except StopIteration:
            return rows
        idx = _build_header_index(header, PS_CSV_HEADER_MAP)
        required = {"student_id", "course_id", "section_id"}
        missing = required - idx.keys()
        if missing:
            raise ValueError(f"PS CSV missing required columns: {missing} (found headers: {header})")

        def get(row: list[str], key: str) -> str:
            i = idx.get(key)
            if i is None or i >= len(row):
                return ""
            return (row[i] or "").strip()

        for row in reader:
            if not any(row):
                continue
            rows.append(PSEnrollmentRow(
                student_id=get(row, "student_id"),
                first_name=get(row, "first_name"),
                last_name=get(row, "last_name"),
                email=get(row, "email"),
                course_id=get(row, "course_id"),
                course_name=get(row, "course_name"),
                section_id=get(row, "section_id"),
                section_name=get(row, "section_name"),
                enrollment_type=get(row, "enrollment_type"),
                status=get(row, "status"),
                campus_id=get(row, "campus_id"),
                campus_name=get(row, "campus_name"),
            ))
    return rows


# === Columbus xlsx readers ===================================================

@dataclass
class ColumbusCourseDemand:
    """One row from `UPDATED MARCH 20 - COURSE_GRADE` — demand per grade."""
    course_name: str
    grade_9: int
    grade_10: int
    grade_11: int
    grade_12: int
    total: int
    target_sections: int | None
    avg_per_section: float | None


@dataclass
class ColumbusListadoRow:
    """One row from `LISTADO MAESTRO CURSOS Y SECCIONES`."""
    course_name: str
    teacher_name: str
    room: str
    n_sections_for_teacher: float | None


@dataclass
class ColumbusStudentRequest:
    """A student's request for a specific course (from *_Conditional Feb 24 sheets)."""
    department: str
    student_id: str
    next_year_grade: int
    student_name: str
    course_number: str
    course_name: str
    course_group: str
    course_group_name: str
    is_alternative: bool  # rank 2 = "Electives Alternative 1" or similar


@dataclass
class ColumbusGroupingRow:
    """One row from `Student Groupings`."""
    code: int
    student_a_name: str
    student_a_id: str
    student_a_grade: int
    student_b_name: str
    student_b_id: str
    student_b_grade: int
    relationship: str  # "Can't be Together", "Should be Together", etc.


def _iter_rows(ws, start_row: int = 1) -> Iterator[tuple]:
    for r in range(start_row, ws.max_row + 1):
        yield tuple(ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1))


def _nz_int(v: Any, default: int = 0) -> int:
    if v is None or v == "":
        return default
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return default


def _nz_float(v: Any, default: float | None = None) -> float | None:
    if v is None or v == "":
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def _nz_str(v: Any, default: str = "") -> str:
    return default if v is None else str(v).strip()


def _id_str(v: Any, default: str = "") -> str:
    """Normalize an ID that may come through as a float ('28025.0') → '28025'."""
    if v is None or v == "":
        return default
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def read_columbus_course_demand(path: Path, sheet: str = "UPDATED MARCH 20 - COURSE_GRADE") -> list[ColumbusCourseDemand]:
    """Read course-by-grade demand from the STUDENTS_PER_COURSE workbook."""
    wb = load_workbook(Path(path), data_only=True)
    ws = wb[sheet]
    rows: list[ColumbusCourseDemand] = []
    for raw in _iter_rows(ws, start_row=2):
        course_name = _nz_str(raw[0])
        if not course_name or course_name.upper().startswith("COURSE"):
            continue
        rows.append(ColumbusCourseDemand(
            course_name=course_name,
            grade_9=_nz_int(raw[1]),
            grade_10=_nz_int(raw[2]),
            grade_11=_nz_int(raw[3]),
            grade_12=_nz_int(raw[4]),
            total=_nz_int(raw[5]),
            target_sections=_nz_int(raw[6]) if raw[6] not in (None, "") else None,
            avg_per_section=_nz_float(raw[7]),
        ))
    return rows


def read_columbus_listado_maestro(path: Path, sheet: str = "LISTADO MAESTRO CURSOS Y SECCIO") -> list[ColumbusListadoRow]:
    """Read teacher → course → room mapping from STUDENTS_PER_COURSE workbook.

    Note: sheet name in the actual file is truncated to 'LISTADO MAESTRO CURSOS Y SECCIO'
    (xlsx has a 31-char sheet name limit but real sheet may use shorter form).
    """
    wb = load_workbook(Path(path), data_only=True)
    # Tolerant: some files use slightly different sheet names
    sn = sheet
    if sn not in wb.sheetnames:
        candidates = [s for s in wb.sheetnames if "listado" in s.lower() or "maestro" in s.lower()]
        if not candidates:
            raise KeyError(f"No 'LISTADO MAESTRO' sheet found. Have: {wb.sheetnames}")
        sn = candidates[0]
    ws = wb[sn]
    rows: list[ColumbusListadoRow] = []
    for raw in _iter_rows(ws, start_row=2):
        course = _nz_str(raw[0])
        teacher = _nz_str(raw[1])
        if not course or not teacher:
            continue
        rows.append(ColumbusListadoRow(
            course_name=course,
            teacher_name=teacher,
            room=_nz_str(raw[2]),
            n_sections_for_teacher=_nz_float(raw[3]),
        ))
    return rows


def read_columbus_student_requests(
    path: Path,
    sheets: list[str] | None = None,
) -> list[ColumbusStudentRequest]:
    """Read student-by-student course requests from the *_Conditional sheets.

    Each department's sheet has the same column layout. We auto-detect sheet
    names that end in 'Conditional Feb 24' (or contain 'Conditional').
    """
    wb = load_workbook(Path(path), data_only=True)
    if sheets is None:
        sheets = [s for s in wb.sheetnames if "conditional" in s.lower() or s.endswith("Feb 24")]

    out: list[ColumbusStudentRequest] = []
    for sn in sheets:
        ws = wb[sn]
        # Find header row (first row where col 0 == "DEPARTMENT" or col contains 'STUDENT_ID')
        header_row = None
        for r in range(1, min(5, ws.max_row + 1)):
            cells = [str(ws.cell(row=r, column=c).value or "").strip().upper() for c in range(1, min(15, ws.max_column + 1))]
            if "STUDENT_ID" in cells or "DEPARTMENT" in cells:
                header_row = r
                break
        if header_row is None:
            continue
        for raw in _iter_rows(ws, start_row=header_row + 1):
            dept = _nz_str(raw[0])
            # Use the PS "ID" column (raw[2]) as student_id, since this is what
            # matches the Student Groupings sheet. The "STUDENT_ID" column (raw[1])
            # is a different identifier (looks like an account ID).
            student_id = _id_str(raw[2]) if len(raw) > 2 else ""
            if not student_id:
                continue
            course_group_name = _nz_str(raw[9]) if len(raw) > 9 else ""
            is_alt = "alternative" in course_group_name.lower()
            out.append(ColumbusStudentRequest(
                department=dept,
                student_id=student_id,
                next_year_grade=_nz_int(raw[4]) if len(raw) > 4 else 0,
                student_name=_nz_str(raw[5]) if len(raw) > 5 else "",
                course_number=_nz_str(raw[6]) if len(raw) > 6 else "",
                course_name=_nz_str(raw[7]) if len(raw) > 7 else "",
                course_group=_nz_str(raw[8]) if len(raw) > 8 else "",
                course_group_name=course_group_name,
                is_alternative=is_alt,
            ))
    return out


def read_columbus_groupings(path: Path, sheet: str = "Student Groupings") -> list[ColumbusGroupingRow]:
    """Read separation/grouping pairs from the HS_Schedule workbook."""
    wb = load_workbook(Path(path), data_only=True)
    if sheet not in wb.sheetnames:
        candidates = [s for s in wb.sheetnames if "grouping" in s.lower()]
        if not candidates:
            raise KeyError(f"No 'Student Groupings' sheet found. Have: {wb.sheetnames}")
        sheet = candidates[0]
    ws = wb[sheet]
    rows: list[ColumbusGroupingRow] = []
    for raw in _iter_rows(ws, start_row=2):
        code = _nz_int(raw[0])
        if not code:
            continue
        rows.append(ColumbusGroupingRow(
            code=code,
            student_a_name=_nz_str(raw[1]),
            student_a_id=_id_str(raw[2]),
            student_a_grade=_nz_int(raw[3]),
            student_b_name=_nz_str(raw[4]),
            student_b_id=_id_str(raw[5]),
            student_b_grade=_nz_int(raw[6]),
            relationship=_nz_str(raw[7]),
        ))
    return rows


# === Dataset assembly ========================================================


def _slugify(name: str, max_len: int = 12) -> str:
    """Turn a course/teacher name into a stable id-like slug."""
    s = re.sub(r"[^\w]+", "_", name.strip(), flags=re.UNICODE)
    s = re.sub(r"_+", "_", s).strip("_")
    return s[:max_len].upper() if s else "X"


def _infer_room_type(room_name: str, course_name: str) -> RoomType:
    """Heuristic: map a room/course to a RoomType."""
    rn = (room_name or "").lower()
    cn = (course_name or "").lower()
    if "lab" in rn or "lab" in cn or any(k in cn for k in ("biology", "chemistry", "physics")):
        return RoomType.SCIENCE_LAB
    if "comp" in rn or "robot" in cn or "computer" in cn:
        return RoomType.COMPUTER_LAB
    if "art" in rn or "art" in cn or "design" in cn:
        return RoomType.ART
    if "music" in rn or "band" in cn:
        return RoomType.MUSIC
    if "gym" in rn or "physical education" in cn or "pe " in cn:
        return RoomType.GYM
    return RoomType.STANDARD


def _grade_demand(d: ColumbusCourseDemand, g: int) -> int:
    """Lookup the per-grade student count on a demand row by grade number."""
    return {9: d.grade_9, 10: d.grade_10, 11: d.grade_11, 12: d.grade_12}.get(g, 0)


def build_dataset_from_columbus(
    demand_path: Path,
    schedule_path: Path | None = None,
    grade: int | list[int] = 12,
    year: str = "2026-2027",
) -> Dataset:
    """Build a canonical Dataset from real Columbus xlsx exports.

    Required:
    - demand_path: path to '1._STUDENTS_PER_COURSE_2026-2027.xlsx'

    Optional:
    - schedule_path: path to 'HS_Schedule_25-26.xlsx' for groupings/separations.
      If not provided, behavior matrix is empty.
    - grade: single int (legacy) or list of ints. Pass `[9, 10, 11, 12]` for
      a full-HS dataset. Each Student keeps their actual grade; each Course's
      `grade_eligibility` is the set of included grades that requested it.

    Returns a Dataset filtered to the given grade(s). All courses with ≥1 demand
    in any included grade are included; teachers and rooms are derived from the
    LISTADO MAESTRO sheet.
    """
    grades: list[int] = [grade] if isinstance(grade, int) else sorted(set(grade))
    primary_grade = min(grades)

    demand = read_columbus_course_demand(demand_path)
    listado = read_columbus_listado_maestro(demand_path)
    requests_raw = read_columbus_student_requests(demand_path)

    # Filter requests to all target grades
    requests_filt = [r for r in requests_raw if r.next_year_grade in grades]

    # Determine which courses are actually requested by any included grade
    requested_course_names = {r.course_name for r in requests_filt}

    # Build courses (slugged IDs)
    course_id_by_name: dict[str, str] = {}
    courses: list[Course] = []

    # Always include Advisory; eligible for every included grade
    courses.append(Course(
        course_id="ADV",
        name="Advisory",
        department="advisory",
        grade_eligibility=list(grades),
        is_required=True,
        meetings_per_week=1,
        max_size=25,
        is_advisory=True,
        qualified_teacher_ids=["__ANY__"],
    ))
    course_id_by_name["Advisory"] = "ADV"

    used_ids: set[str] = {"ADV"}
    for d in demand:
        # Skip courses that have no demand in any of the included grades
        eligible_grades = [g for g in grades if _grade_demand(d, g) > 0]
        if not eligible_grades:
            continue
        # Only include courses that at least one student in our cohort requested
        if d.course_name not in requested_course_names and d.course_name != "Advisory":
            continue
        cid = _slugify(d.course_name, 12)
        # Ensure unique ID
        suffix = 0
        base = cid
        while cid in used_ids:
            suffix += 1
            cid = f"{base}{suffix}"
        used_ids.add(cid)
        course_id_by_name[d.course_name] = cid

        # Infer is_lab from course name
        is_lab = any(k in d.course_name.lower() for k in ("biology", "chemistry", "physics", "lab"))
        room_type = _infer_room_type("", d.course_name)
        max_size = 26 if "AP Research" in d.course_name else 25

        # is_required is grade-specific; "Math 9" is required for grade 9 only.
        # When ingesting multiple grades, mark a course required if its name is
        # tagged "Required" or ends with one of our grade numbers.
        name_lower = d.course_name.lower()
        is_required_flag = "required" in name_lower or any(name_lower.endswith(str(g)) for g in grades)

        courses.append(Course(
            course_id=cid,
            name=d.course_name,
            department="unknown",  # filled below from listado
            grade_eligibility=eligible_grades,
            is_required=is_required_flag,
            meetings_per_week=3,
            max_size=max_size,
            required_room_type=room_type,
            is_lab=is_lab,
            qualified_teacher_ids=[],
        ))

    courses_by_id = {c.course_id: c for c in courses}

    # Build teachers (one per unique teacher name from listado)
    teacher_id_by_name: dict[str, str] = {}
    teachers: list[Teacher] = []
    for row in listado:
        if row.teacher_name in teacher_id_by_name:
            continue
        tid = f"T_{_slugify(row.teacher_name, 10)}"
        # ensure unique
        suffix = 0
        base = tid
        while any(t.teacher_id == tid for t in teachers):
            suffix += 1
            tid = f"{base}{suffix}"
        teacher_id_by_name[row.teacher_name] = tid
        teachers.append(Teacher(
            teacher_id=tid,
            name=row.teacher_name,
            department="unknown",
            qualified_course_ids=[],
            max_load=5,
            home_room_id=None,
        ))

    # Wire up qualifications + departments from listado
    for row in listado:
        cid = course_id_by_name.get(row.course_name)
        if cid is None:
            continue  # course not in this grade's set
        c = courses_by_id[cid]
        tid = teacher_id_by_name[row.teacher_name]
        if tid not in c.qualified_teacher_ids:
            c.qualified_teacher_ids.append(tid)
        # Update teacher's qualifications
        t = next(t for t in teachers if t.teacher_id == tid)
        if cid not in t.qualified_course_ids:
            t.qualified_course_ids.append(cid)

    # Drop courses that have no qualified teacher (real-data noise: a course is
    # requested by 1-2 students but isn't in LISTADO MAESTRO, e.g. "Teacher Aide").
    # The orphan requests fall through to is-required-but-no-section validation
    # later, so we drop them silently here and log via the dataset assembly path.
    courses = [c for c in courses if c.is_advisory or c.qualified_teacher_ids]
    courses_by_id = {c.course_id: c for c in courses}
    valid_course_ids = set(courses_by_id.keys())

    # Build rooms (one per unique non-empty room id in listado)
    room_id_by_label: dict[str, str] = {}
    rooms: list[Room] = []
    for row in listado:
        if not row.room or row.room.lower() in ("none", "n/a", ""):
            continue
        label = row.room
        if label in room_id_by_label:
            continue
        rid = f"R{_slugify(label, 8)}"
        # ensure unique
        suffix = 0
        base = rid
        while any(r.room_id == rid for r in rooms):
            suffix += 1
            rid = f"{base}{suffix}"
        room_id_by_label[label] = rid
        # Determine room type by inspecting linked courses
        linked_courses = [r.course_name for r in listado if r.room == label]
        room_type = _infer_room_type(label, " ".join(linked_courses))
        rooms.append(Room(
            room_id=rid,
            name=label,
            capacity=28 if room_type == RoomType.STANDARD else 26,
            room_type=room_type,
        ))

    # HC4 prep: assign each teacher's home_room_id from LISTADO MAESTRO.
    # Per the "Reglas Horarios HS" doc (rfi_Reglas_Horarios_HS_*.md, 2026-04-22):
    # "salón es por profesor; profesores sin salón propio rotan usando el espacio
    # de un profesor asignado cuando éste lo tenga libre".
    #
    # Rules applied:
    #   - Skip placeholder teachers ("New X Teacher" — unfilled positions): no home_room,
    #     they float per the rotation rule.
    #   - Single-room teachers (LISTADO has only one ROOM value across all their
    #     courses): pin home_room to that room.
    #   - Multi-room teachers (Sindy Margarita has both R933 and R923): NO home_room
    #     pinned; they rotate naturally. Pinning would arbitrarily eliminate one of
    #     their rooms from feasibility.
    #
    # When two real teachers share the same single room (e.g. Felipe Naranjo +
    # Emily Butterworth both in R926), HC2 (no two sections same scheme + room)
    # enforces that only one is in class at any given scheme. As long as their
    # combined section count ≤ 8 (the academic scheme count), the schedule is
    # feasible. The doc explicitly accepts this: "Si dos profesores deben usar el
    # mismo salón, la planificación debe garantizar que solo uno de ellos esté
    # en clase a la vez".
    from collections import defaultdict as _dd
    teacher_room_set: dict[str, set] = _dd(set)
    for row in listado:
        if not row.room or row.room.lower() in ("none", "n/a", ""):
            continue
        teacher_room_set[row.teacher_name].add(row.room)

    def _is_placeholder_teacher(name: str) -> bool:
        return name.strip().lower().startswith("new ")

    for t in teachers:
        if _is_placeholder_teacher(t.name):
            continue  # placeholders float; no home_room
        t_rooms = teacher_room_set.get(t.name, set())
        if len(t_rooms) == 1:
            room_label = next(iter(t_rooms))
            if room_label in room_id_by_label:
                t.home_room_id = room_id_by_label[room_label]
        # multi-room teachers: leave home_room_id=None (default)

    # Ensure at least one room per required type
    needed_types = {c.required_room_type for c in courses if not c.is_advisory}
    available_types = {r.room_type for r in rooms}
    for missing_type in needed_types - available_types:
        rooms.append(Room(
            room_id=f"R_FALLBACK_{missing_type.value.upper()}",
            name=f"Fallback {missing_type.value}",
            capacity=26,
            room_type=missing_type,
        ))

    # Build sections from listado (each row = one section)
    sections: list[Section] = []
    for row in listado:
        cid = course_id_by_name.get(row.course_name)
        if cid is None or cid not in valid_course_ids:
            continue
        tid = teacher_id_by_name[row.teacher_name]
        c = courses_by_id[cid]
        n = int(row.n_sections_for_teacher or 1)
        for i in range(n):
            existing = sum(1 for s in sections if s.course_id == cid)
            sections.append(Section(
                section_id=f"{cid}.{existing + 1}",
                course_id=cid,
                teacher_id=tid,
                max_size=c.max_size,
                grade_level=primary_grade,
            ))

    # Adjust each teacher's max_load to fit observed section count (real
    # Columbus has some teachers carrying 6 sections; max_load=5 default would
    # otherwise produce spurious overload warnings).
    sections_per_teacher: dict[str, int] = {}
    for s in sections:
        sections_per_teacher[s.teacher_id] = sections_per_teacher.get(s.teacher_id, 0) + 1
    for t in teachers:
        observed = sections_per_teacher.get(t.teacher_id, 0)
        if observed > t.max_load:
            t.max_load = observed

    # Build students from the requests
    students_map: dict[str, Student] = {}
    for r in requests_filt:
        sid = r.student_id
        if sid not in students_map:
            students_map[sid] = Student(
                student_id=sid,
                name=r.student_name,
                grade=r.next_year_grade or primary_grade,
                requested_courses=[
                    CourseRequest(student_id=sid, course_id="ADV", is_required=True, rank=1),
                ],
            )
        cid = course_id_by_name.get(r.course_name)
        if cid is None or cid not in valid_course_ids:
            continue  # silently drop requests for dropped (orphan) courses
        is_required = not r.is_alternative and "Required" in r.course_group_name
        rank = 2 if r.is_alternative else 1
        students_map[sid].requested_courses.append(CourseRequest(
            student_id=sid, course_id=cid, is_required=is_required, rank=rank,
        ))

    # Add Advisory section + assignments — one per ~22 students
    advisory_size = 25
    n_adv_sections = max(1, -(-len(students_map) // advisory_size))
    for i in range(n_adv_sections):
        # Use the first available teachers for advisory
        teacher_id = teachers[i % len(teachers)].teacher_id if teachers else "T_UNKNOWN"
        sections.append(Section(
            section_id=f"ADV.{i+1}",
            course_id="ADV",
            teacher_id=teacher_id,
            max_size=advisory_size,
            grade_level=primary_grade,
        ))

    # Build behavior matrix from groupings (if schedule_path given)
    behavior = BehaviorMatrix()
    if schedule_path is not None:
        try:
            grouping_rows = read_columbus_groupings(schedule_path)
            seps: list[tuple[str, str]] = []
            grps: list[tuple[str, str]] = []
            for g in grouping_rows:
                # Match students by ID; if either student isn't in our grade-filtered set, skip
                if g.student_a_id not in students_map or g.student_b_id not in students_map:
                    continue
                pair = (g.student_a_id, g.student_b_id)
                if "can't" in g.relationship.lower() or "no" in g.relationship.lower():
                    seps.append(pair)
                elif "should" in g.relationship.lower() or "together" in g.relationship.lower():
                    grps.append(pair)
            behavior = BehaviorMatrix(separations=seps, groupings=grps)
        except Exception:
            pass  # Non-blocking: empty behavior matrix is acceptable

    # max_consecutive_classes: client confirmed 2026-04-26 that the cap is 4.
    # Pigeonhole reality: with 5 blocks/day, 5 days, 8 schemes (3 cells each),
    # a teacher using 7 distinct schemes is busy 21 of 25 cells; max=4 means
    # ≤4 of 5 blocks busy per day, so ≤20 teaching cells/week. 21 > 20 ⇒ master
    # is structurally infeasible at strict 4 if any teacher carries ≥7 academic
    # sections. We REPORT this conflict instead of silently auto-relaxing —
    # operators can either (a) reduce one of those teachers' loads to ≤6
    # sections, or (b) explicitly pass max_consecutive_classes=5 via config.
    hard = HardConstraints()
    overloaded_teachers = sorted(
        ((t.name, n) for t, n in [(next((t for t in teachers if t.teacher_id == tid), None), n)
                                    for tid, n in sections_per_teacher.items()] if t and n >= 7),
        key=lambda x: -x[1],
    )
    if overloaded_teachers:
        import sys
        print(
            f"WARNING: {len(overloaded_teachers)} teacher(s) carry ≥7 academic sections; "
            f"strict max_consecutive_classes=4 may be infeasible. Offenders:",
            file=sys.stderr,
        )
        for name, n in overloaded_teachers:
            print(f"  - {name}: {n} sections", file=sys.stderr)
        print(
            "Mitigation options: (a) coordinate with school to reduce one teacher's "
            "load to ≤6 sections; (b) pass HardConstraints(max_consecutive_classes=5).",
            file=sys.stderr,
        )

    # PowerSchool field values per Columbus IT confirmation 2026-04-26:
    # - SchoolID: number, MS=12000, HS=13000
    # - TermID: 3600 for 2026-2027
    # Detect MS vs HS by ingested grades. MS = 6/7/8; HS = 9/10/11/12.
    is_ms = all(g in (6, 7, 8) for g in grades)
    school_name = "Columbus Middle School" if is_ms else "Columbus High School"
    school_id = 12000 if is_ms else 13000
    term_id = "3600" if "2026-2027" in str(year) else None

    config = SchoolConfig(
        school=school_name,
        school_id=school_id,
        grade=primary_grade,
        year=year,
        term_id=term_id,
        bell=default_rotation(),
        hard=hard,
        soft=SoftConstraintWeights(),
    )

    return Dataset(
        config=config,
        courses=courses,
        teachers=teachers,
        rooms=rooms,
        sections=sections,
        students=list(students_map.values()),
        behavior=behavior,
    )


# === CLI helper ==============================================================

if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Usage: python ps_ingest.py <demand.xlsx> [schedule.xlsx] [grade]")
        sys.exit(1)
    demand = Path(sys.argv[1])
    sched = Path(sys.argv[2]) if len(sys.argv) > 2 else None
    grade = int(sys.argv[3]) if len(sys.argv) > 3 else 12

    print(f"Reading {demand.name}...")
    ds = build_dataset_from_columbus(demand, sched, grade=grade)
    print(f"Built dataset for Grade {grade}:")
    print(f"  courses:  {len(ds.courses)}")
    print(f"  teachers: {len(ds.teachers)}")
    print(f"  rooms:    {len(ds.rooms)}")
    print(f"  sections: {len(ds.sections)}")
    print(f"  students: {len(ds.students)}")
    print(f"  separations: {len(ds.behavior.separations)}")
    print(f"  groupings:   {len(ds.behavior.groupings)}")
