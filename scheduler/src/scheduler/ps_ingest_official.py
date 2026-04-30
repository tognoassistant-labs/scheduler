"""Canonical PowerSchool ingester (replaces xlsx-derived ps_ingest.py for v4+).

Reads the official Columbus PowerSchool catalog xlsx exported from PS itself
(`reference/columbus_official_2026-2027.xlsx`) and produces a Dataset with
**real PS IDs** — no slugs, no heuristics.

5 sheets:
    courses              — Course catalog with COURSE_NUMBER (real PS), MAXCLASSSIZE,
                            MULTITERM, SCHED_FREQUENCY, SCHED_DEMAND
    rooms                — Room catalog with DCID, ROOMNUMBER, DEPARTMENT, MAXIMUM
    teachers             — Teacher catalog with DCID, LASTFIRST, PREFERRED_ROOM,
                            SCHED_DEPARTMENT
    teacher_assignments  — Per-(teacher, course) assignments with SECTIONS_PER_COURSE
    requests             — Per-student requests with COURSENUMBER, STUDENT_NUMBER

Fixes applied (consolidated from ps_ingest.py history):
    - All requests are is_required=True / rank=1 (client B6: no alternates 2026-2027)
    - PREFERRED_ROOM → Teacher.home_room_id (HC4)
    - max_consecutive_classes per teacher: default 4, override 5 only for the
      pigeonhole-impossible cases (≥7 academic sections)
    - Multi-room teachers (PREFERRED_ROOM blank in source) stay floating
    - Semester courses (MULTITERM='S1' or 'S2') are flagged via Course.term
"""
from __future__ import annotations

import csv
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

from .models import (
    BehaviorMatrix,
    Course,
    CourseRequest,
    Dataset,
    HardConstraints,
    Room,
    RoomType,
    SchoolConfig,
    Section,
    SoftConstraintWeights,
    Student,
    Teacher,
    Term,
    default_rotation,
)


# ---------------------------------------------------------------------------
# Helpers


def _safe_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _read_coplanning_groups(
    xlsx_path: Path,
    teacher_by_lastfirst: dict[str, "Teacher"],
) -> list[list[str]]:
    """Parse the `CO PLANNING INFO` sheet of the RFI xlsx.

    Each blank-row-separated block lists 2-3 teachers that must share a free
    scheme (preparation period). Returns groups as `[[teacher_dcid, ...], ...]`,
    dropping placeholder names ("New X Teacher") not present in the dataset and
    discarding groups that end up with <2 known teachers.
    """
    if not xlsx_path.exists():
        return []
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if "CO PLANNING INFO" not in wb.sheetnames:
        return []
    ws = wb["CO PLANNING INFO"]

    groups: list[list[str]] = []
    current: list[str] = []
    for row in ws.iter_rows(values_only=True):
        teacher = _safe_str((row + (None,))[0]) if row else ""
        if not teacher:
            if current:
                groups.append(current)
                current = []
            continue
        if teacher.upper().startswith(("CO PLANNING", "PRIORIDADES", "OTROS")):
            continue
        t = teacher_by_lastfirst.get(teacher)
        if t is None:
            continue  # placeholder ("New X Teacher") or rename mismatch
        if t.teacher_id not in current:
            current.append(t.teacher_id)
    if current:
        groups.append(current)

    return [g for g in groups if len(g) >= 2]


def _safe_int(v, default: int = 0) -> int:
    if v is None or v == "":
        return default
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return default


def _infer_room_type(department: str, room_name: str) -> RoomType:
    """Map a room department + name to a RoomType."""
    d = (department or "").lower()
    n = (room_name or "").lower()
    if "pe" in d or "coliseo" in n or "gym" in n:
        return RoomType.GYM
    if "science" in d or "lab" in n:
        return RoomType.SCIENCE_LAB
    if "tech" in d or "maker" in n or "comp" in n:
        return RoomType.COMPUTER_LAB
    if "art" in d:
        return RoomType.ART
    if "music" in d or "band" in n:
        return RoomType.MUSIC
    return RoomType.STANDARD


def _infer_grades_from_name(name: str) -> list[int]:
    """Extract grade levels from a course name like 'English 9' or 'AP Spanish Lit'.

    Returns the grades the course is offered to. For courses without an explicit
    grade in the name, returns [9, 10, 11, 12] as the eligibility set.
    """
    grades = sorted({int(m) for m in re.findall(r"\b(9|10|11|12)\b", name or "")})
    if grades:
        return grades
    # No explicit grade in name → eligible to all HS grades
    return [9, 10, 11, 12]


# ---------------------------------------------------------------------------
# Sheet readers


def _read_courses(wb) -> list[dict]:
    ws = wb["courses"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    out = []
    for raw in rows[1:]:
        if not raw or not raw[0]:
            continue
        d = dict(zip(headers, raw))
        if not d.get("COURSE_NUMBER"):
            continue
        out.append(d)
    return out


def _read_rooms(wb) -> list[dict]:
    ws = wb["rooms"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    out = []
    for raw in rows[1:]:
        if not raw or not raw[0]:
            continue
        d = dict(zip(headers, raw))
        if not d.get("ROOMNUMBER"):
            continue
        out.append(d)
    return out


def _read_teachers(wb) -> list[dict]:
    ws = wb["teachers"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    out = []
    for raw in rows[1:]:
        if not raw or not raw[0]:
            continue
        d = dict(zip(headers, raw))
        if not d.get("LASTFIRST"):
            continue
        out.append(d)
    return out


def _read_teacher_assignments(wb) -> list[dict]:
    ws = wb["teacher_assignments"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    out = []
    for raw in rows[1:]:
        if not raw or not raw[0]:
            continue
        d = dict(zip(headers, raw))
        if not d.get("COURSENUMBER") or not d.get("TEACHER_DCID"):
            continue
        out.append(d)
    return out


def _read_relationships_xlsx(wb) -> list[tuple[str, str, str]]:
    """Read the in-file `course_relationships` tab. Returns [(c1, c2, code), …]."""
    if "course_relationships" not in wb.sheetnames:
        return []
    ws = wb["course_relationships"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = list(rows[0])
    out: list[tuple[str, str, str]] = []
    for r in rows[1:]:
        if not r or not any(c is not None for c in r):
            continue
        d = dict(zip(headers, r))
        c1 = _safe_str(d.get("COURSE_NUMBER1"))
        c2 = _safe_str(d.get("COURSE_NUMBER2"))
        code = _safe_str(d.get("RELATIONSHIPCODE"))
        if c1 and c2 and code:
            out.append((c1, c2, code))
    return out


def _read_relationships_csv(rel_path: Path) -> list[tuple[str, str, str]]:
    """Read sibling course_relationships.csv (legacy). Returns [(c1, c2, code), …]."""
    if not rel_path.exists():
        return []
    out: list[tuple[str, str, str]] = []
    with rel_path.open() as f:
        for row in csv.DictReader(f):
            c1 = (row.get("COURSE_NUMBER1") or "").strip()
            c2 = (row.get("COURSE_NUMBER2") or "").strip()
            code = (row.get("RELATIONSHIPCODE") or "").strip()
            if c1 and c2 and code:
                out.append((c1, c2, code))
    return out


def _apply_course_relationships(
    relationships: list[tuple[str, str, str]],
    courses: list[Course],
    course_by_number: dict[str, Course],
) -> None:
    """Apply pre-read course relationships to set Course.simul_group / Course.term_pair.

    Simultaneous relationships are unioned via union-find so chains (G0902 ↔
    G1204, G0902 ↔ G1205, G0902 ↔ G1206) collapse to a single shared group ID.
    Term relationships are stored as a single peer pointer per course.
    """
    parent: dict[str, str] = {}

    def _find(x: str) -> str:
        while parent.get(x, x) != x:
            parent[x] = parent.get(parent[x], parent[x])
            x = parent[x]
        return x

    def _union(a: str, b: str) -> None:
        parent.setdefault(a, a)
        parent.setdefault(b, b)
        ra, rb = _find(a), _find(b)
        if ra != rb:
            parent[ra] = rb

    term_pairs: dict[str, str] = {}
    simul_courses: set[str] = set()
    for c1, c2, code in relationships:
        if code == "Simultaneous":
            _union(c1, c2)
            simul_courses.add(c1)
            simul_courses.add(c2)
        elif code == "Term":
            term_pairs[c1] = c2
            term_pairs[c2] = c1

    for cid in simul_courses:
        c = course_by_number.get(cid)
        if c is not None:
            c.simul_group = _find(cid)
    for cid, peer in term_pairs.items():
        c = course_by_number.get(cid)
        if c is not None:
            c.term_pair = peer


# Mapping for co-planning placeholder teacher names (display-shorter forms)
# to the canonical LASTFIRST records in the teachers sheet.
_PLACEHOLDER_TEACHER_NAME_MAP = {
    "New Science Teacher": "Science Teacher 1, New",
    "New English Teacher": "English Teacher 1, New",
}


def _read_coplanning_groups_xlsx(
    wb,
    teacher_by_lastfirst: dict[str, "Teacher"],
) -> list[list[str]]:
    """Parse the in-file `co-planning` tab.

    Each row is (CO PLANNING TEACHERS, COURSE ID, COURSES, PRIORITY). Teachers
    that share the same COURSE ID form one coplanning group. Placeholder names
    like "New Science Teacher" are normalized to the canonical LASTFIRST.
    """
    if "co-planning" not in wb.sheetnames:
        return []
    ws = wb["co-planning"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = list(rows[0])
    teacher_col = next((i for i, h in enumerate(headers) if h and "TEACHER" in str(h).upper()), 0)
    course_col = next((i for i, h in enumerate(headers) if h and "COURSE" in str(h).upper() and "ID" in str(h).upper()), 1)

    by_course: dict[str, list[str]] = defaultdict(list)
    unmatched: set[str] = set()
    for r in rows[1:]:
        if not r or not any(c is not None for c in r):
            continue
        teacher_name = _safe_str(r[teacher_col]) if len(r) > teacher_col else ""
        course_id = _safe_str(r[course_col]) if len(r) > course_col else ""
        if not teacher_name or not course_id:
            continue
        canonical_name = _PLACEHOLDER_TEACHER_NAME_MAP.get(teacher_name, teacher_name)
        teacher = teacher_by_lastfirst.get(canonical_name)
        if teacher is None:
            unmatched.add(teacher_name)
            continue
        if teacher.teacher_id not in by_course[course_id]:
            by_course[course_id].append(teacher.teacher_id)
    if unmatched:
        print(f"[WARN] co-planning: {len(unmatched)} unmatched teacher names:")
        for n in sorted(unmatched)[:10]:
            print(f"    {n!r}")
    return [g for g in by_course.values() if len(g) >= 2]


_BEHAVIOR_SEPARATION_TYPES = {"separado de", "separada de", "separado", "separada"}
_BEHAVIOR_GROUPING_TYPES = {"compartir clases con", "compartir con"}


def _read_behavior_matrix_xlsx(
    wb,
    valid_student_ids: set[str],
) -> tuple[BehaviorMatrix, dict[str, str]]:
    """Parse `conselours_recommendations` → (BehaviorMatrix, name_to_id).

    The name_to_id mapping (NOMBRE.upper() → STUDENT_NUMBER) is returned so
    that other sheets (e.g. teacher_avoid, where STUDENT_NUMBER may be blank)
    can resolve students by name.
    """
    if "conselours_recommendations" not in wb.sheetnames:
        return BehaviorMatrix(), {}
    ws = wb["conselours_recommendations"]
    rows = list(ws.iter_rows(values_only=True))
    seps: list[tuple[str, str]] = []
    grps: list[tuple[str, str]] = []
    name_to_id: dict[str, str] = {}
    skipped_unknown_students = 0
    skipped_unknown_type = 0
    for r in rows[1:]:
        if not r or not any(c is not None for c in r):
            continue
        # cols: GRADE_LEVEL | CODIGO | NOMBRE | TYPE | CODIGO | NOMBRE
        s1_id = _safe_str(r[1]) if len(r) > 1 else ""
        s1_name = _safe_str(r[2]).upper() if len(r) > 2 else ""
        rel_type = _safe_str(r[3]).lower() if len(r) > 3 else ""
        s2_id = _safe_str(r[4]) if len(r) > 4 else ""
        s2_name = _safe_str(r[5]).upper() if len(r) > 5 else ""
        if s1_name and s1_id:
            name_to_id[s1_name] = s1_id
        if s2_name and s2_id:
            name_to_id[s2_name] = s2_id
        if not (s1_id and s2_id and rel_type):
            continue
        if s1_id not in valid_student_ids or s2_id not in valid_student_ids:
            skipped_unknown_students += 1
            continue
        pair = (s1_id, s2_id)
        if rel_type in _BEHAVIOR_SEPARATION_TYPES:
            seps.append(pair)
        elif rel_type in _BEHAVIOR_GROUPING_TYPES:
            grps.append(pair)
        else:
            skipped_unknown_type += 1
    if skipped_unknown_students:
        print(f"[WARN] conselours_recommendations: {skipped_unknown_students} rows with student IDs not in requests sheet")
    if skipped_unknown_type:
        print(f"[WARN] conselours_recommendations: {skipped_unknown_type} rows with unknown relationship type")
    return BehaviorMatrix(separations=seps, groupings=grps), name_to_id


def _read_teacher_avoid_xlsx(
    wb,
    students_map: dict[str, "Student"],
    teacher_by_lastfirst: dict[str, "Teacher"],
    name_to_id: dict[str, str],
) -> int:
    """Apply teacher_avoid restrictions to Student.restricted_teacher_ids.

    The teacher_avoid sheet has STUDENT_NUMBER + STUDENT (name) + TEACHER_NAME.
    When STUDENT_NUMBER is blank, fall back to looking up the student by
    name (uppercased) using the mapping built from conselours_recommendations.
    """
    if "teacher_avoid" not in wb.sheetnames:
        return 0
    ws = wb["teacher_avoid"]
    rows = list(ws.iter_rows(values_only=True))
    matched = 0
    unmatched_students: list[str] = []
    unmatched_teachers: list[str] = []
    for r in rows[1:]:
        if not r or not any(c is not None for c in r):
            continue
        s_id = _safe_str(r[0]) if r[0] else ""
        s_name = _safe_str(r[1]).upper() if len(r) > 1 and r[1] else ""
        t_name = _safe_str(r[2]) if len(r) > 2 and r[2] else ""
        if not s_id and s_name:
            s_id = name_to_id.get(s_name, "")
        if not s_id:
            unmatched_students.append(s_name or "<empty>")
            continue
        teacher = teacher_by_lastfirst.get(t_name)
        if teacher is None:
            unmatched_teachers.append(t_name)
            continue
        student = students_map.get(s_id)
        if student is None:
            unmatched_students.append(s_name or s_id)
            continue
        if teacher.teacher_id not in student.restricted_teacher_ids:
            student.restricted_teacher_ids.append(teacher.teacher_id)
            matched += 1
    if unmatched_students:
        print(f"[WARN] teacher_avoid: {len(unmatched_students)} unmatched students (need STUDENT_NUMBER or name to be in conselours sheet):")
        for n in unmatched_students[:10]:
            print(f"    {n!r}")
    if unmatched_teachers:
        print(f"[WARN] teacher_avoid: {len(unmatched_teachers)} unmatched teachers:")
        for n in unmatched_teachers[:10]:
            print(f"    {n!r}")
    return matched


# Pattern for auto-enforced max-class-size override:
#   "Numero Maximo de estudiantes es 26", "max class size 26", "Max 26 students"
_MAX_SIZE_REGEX = re.compile(
    r"(?:numero\s+maximo|max(?:imo)?(?:\s+(?:class\s+size|de\s+estudiantes|students))?)"
    r"[^\d]*(\d+)",
    re.IGNORECASE,
)


def _apply_teacher_assignment_constraints(
    assignment_rows: list[dict],
    courses: list[Course],
    course_by_number: dict[str, Course],
) -> None:
    """Parse CONSTRAINTS column on teacher_assignments and apply structured rules.

    Auto-enforced today:
      - "max class size = N" pattern → bump Course.max_size if N > current.

    Conditional / free-text rules ("schedule when Tamir is free", etc.) are
    surfaced as [INFO] only — they require manual interpretation.
    """
    items = [(_safe_str(a.get("TEACHER_DCID")), _safe_str(a.get("LASTFIRST")),
              _safe_str(a.get("COURSENUMBER")), _safe_str(a.get("CONSTRAINTS")))
             for a in assignment_rows
             if _safe_str(a.get("CONSTRAINTS"))]
    if not items:
        return

    auto_applied: list[tuple[str, str, int, int]] = []  # (course, teacher, old_max, new_max)
    free_text: list[tuple[str, str, str, str]] = []
    for dcid, name, course_id, ctext in items:
        m = _MAX_SIZE_REGEX.search(ctext)
        if m:
            new_max = int(m.group(1))
            course = course_by_number.get(course_id)
            if course is not None and new_max > course.max_size:
                auto_applied.append((course_id, name, course.max_size, new_max))
                course.max_size = new_max
                continue
        free_text.append((course_id, name, dcid, ctext))

    if auto_applied:
        print(f"[INFO] CONSTRAINTS auto-applied: {len(auto_applied)} max-size overrides:")
        for cid, tname, old, new in auto_applied:
            print(f"    {cid} ({tname}): max_size {old} → {new}")
    if free_text:
        print(f"[INFO] CONSTRAINTS free-text (not auto-enforced): {len(free_text)} rules:")
        for cid, tname, dcid, ctext in free_text:
            print(f"    {cid} / {tname} ({dcid}): {ctext}")


_TEACHER_AIDE_COURSE_ID = "OZ1333"  # Generic "Teacher Aide" placeholder


def _normalize_course_name_for_match(s: str) -> str:
    """Normalize a course name string for fuzzy matching across the
    teacher_assistants sheet (uses abbreviations like 'P.E 11') and the
    courses catalog ('Physical Education and Health 11')."""
    if not s:
        return ""
    s = s.lower().strip()
    # Common abbreviations used in the school's TA sheet
    s = s.replace(".", " ").replace("  ", " ")
    if s.startswith("p e ") or s.startswith("pe "):
        s = "physical education and health " + s.split(" ", 2)[-1]
    # "Ar of Fiction" typo
    s = s.replace("ar of fiction", "art of fiction")
    return s.strip()


def _read_required_courses_xlsx(wb) -> dict[int, set[str]]:
    """Parse the in-file `required_courses` sheet (added by school 2026-04-30).

    Returns: {grade_level: {course_number, …}} — the official list of
    truly-required courses (graduation-required, per grade). Used by the
    ingester to set CourseRequest.is_required correctly: only requests
    matching (student_grade, course_number) in this map are HARD-required;
    the rest are student-elected (still important, but the solver may drop
    them before dropping a HARD required).
    """
    if "required_courses" not in wb.sheetnames:
        return {}
    ws = wb["required_courses"]
    out: dict[int, set[str]] = defaultdict(set)
    for r in list(ws.iter_rows(values_only=True))[1:]:
        if not r or not r[0]:
            continue
        cnum = _safe_str(r[0])
        try:
            grade = int(r[2]) if r[2] is not None else None
        except (TypeError, ValueError):
            grade = None
        if cnum and grade is not None:
            out[grade].add(cnum)
    return dict(out)


def _read_teacher_assistants_xlsx(wb) -> list[dict]:
    """Parse the in-file `teacher_assistants` sheet.

    Each row identifies a student who is a teacher's aide (= teacher
    assistant; school confirmed 2026-04-30 these terms are equivalent).
    Returns: list of dicts with student_id, grade, assist_course_name,
    teacher_name, status. Used to clean their request list so the engine
    does not waste cycles trying to schedule them into the placeholder
    OZ1333 course or into the section they actually assist.
    """
    if "teacher_assistants" not in wb.sheetnames:
        return []
    ws = wb["teacher_assistants"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    out: list[dict] = []
    for r in rows[1:]:
        if not r or not r[0]:
            continue
        try:
            sid = str(int(r[0]))
        except (TypeError, ValueError):
            sid = _safe_str(r[0])
        out.append({
            "student_id": sid,
            "grade": int(r[1]) if isinstance(r[1], (int, float)) else None,
            "name": _safe_str(r[2]),
            "placeholder_course": _safe_str(r[3]),  # COURSENAME (= "Electives Alternative 1")
            "assist_course_name": _safe_str(r[4]),  # COURSENAME_TO_ASSIST
            "teacher_name": _safe_str(r[5]),
            "status": _safe_str(r[6]) if len(r) > 6 else "",
        })
    return out


def _apply_teacher_assistants_cleanup(
    request_rows: list[dict],
    ta_records: list[dict],
    course_by_number: dict[str, "Course"],
) -> tuple[list[dict], dict[str, dict]]:
    """Remove TA-related entries from a student's request list.

    For each TA student in `ta_records`:
      1. Drop any request row with COURSENUMBER = OZ1333 (Teacher Aide
         placeholder).
      2. If the student also has the "assist course" listed as a real
         request (double-entry case e.g. G12 student requesting PE 11
         because they assist PE 11), drop that too.
      3. Skip students whose status string indicates they're an ex-TA
         going back to a normal class ("Sale de TA").

    Returns:
      (filtered_request_rows, ta_summary_per_student)
      ta_summary_per_student is for the visor / reporting later.
    """
    if not ta_records:
        return request_rows, {}

    # Build a (student_id, course_number) skip set.
    skip_pairs: set[tuple[str, str]] = set()
    ta_summary: dict[str, dict] = {}
    ex_ta_skipped: list[str] = []
    double_entry_count = 0
    oz_only_count = 0

    # Helper: index requests by student for fast lookup
    by_student: dict[str, list[dict]] = defaultdict(list)
    for r in request_rows:
        sid = _safe_str(r.get("STUDENT_NUMBER"))
        if sid:
            by_student[sid].append(r)

    # Course name → course_number index for fuzzy matching
    name_to_number: dict[str, str] = {}
    for c in course_by_number.values():
        if c.name:
            name_to_number[c.name.lower().strip()] = c.course_id

    for ta in ta_records:
        sid = ta["student_id"]
        status_l = (ta.get("status") or "").lower()
        assist_l = (ta.get("assist_course_name") or "").lower().strip()

        # Ex-TAs going back to normal classes are NOT cleaned
        if "sale de ta" in assist_l or "sale de ta" in status_l:
            ex_ta_skipped.append(sid)
            continue

        student_reqs = by_student.get(sid, [])
        has_oz = False
        matched_assist_course: str | None = None
        assist_norm = _normalize_course_name_for_match(ta.get("assist_course_name"))
        for req in student_reqs:
            cnum = _safe_str(req.get("COURSENUMBER"))
            cname_norm = _normalize_course_name_for_match(_safe_str(req.get("COURSENAME")))
            # Drop OZ1333 placeholder
            if cnum == _TEACHER_AIDE_COURSE_ID:
                skip_pairs.add((sid, cnum))
                has_oz = True
            # Drop assist course (double-entry case)
            if assist_norm and cname_norm:
                if assist_norm in cname_norm or cname_norm in assist_norm:
                    skip_pairs.add((sid, cnum))
                    matched_assist_course = cnum

        if has_oz and matched_assist_course:
            double_entry_count += 1
        elif has_oz:
            oz_only_count += 1

        ta_summary[sid] = {
            "name": ta.get("name"),
            "grade": ta.get("grade"),
            "assist_course": ta.get("assist_course_name"),
            "teacher": ta.get("teacher_name"),
            "removed_oz1333": has_oz,
            "removed_assist_course_id": matched_assist_course,
        }

    # Filter
    filtered = [r for r in request_rows
                if (_safe_str(r.get("STUDENT_NUMBER")), _safe_str(r.get("COURSENUMBER")))
                not in skip_pairs]

    n_removed = len(request_rows) - len(filtered)
    print(
        f"[INFO] teacher_assistants cleanup: removed {n_removed} request rows "
        f"({double_entry_count} double-entry, {oz_only_count} OZ1333-only, "
        f"{len(ex_ta_skipped)} ex-TA students skipped)"
    )
    if ex_ta_skipped:
        print(f"  Ex-TAs (no cleanup needed): {ex_ta_skipped}")

    # Detect TAs from the sheet who didn't match any of their request entries
    unmatched_tas = [sid for sid, info in ta_summary.items()
                     if not info["removed_oz1333"] and not info["removed_assist_course_id"]]
    if unmatched_tas:
        print(f"[WARN] {len(unmatched_tas)} TA students from sheet have neither OZ1333 nor assist-course in their requests:")
        for sid in unmatched_tas:
            print(f"    student {sid} (grade {ta_summary[sid].get('grade')}): assists {ta_summary[sid].get('assist_course')!r}")

    return filtered, ta_summary


def _audit_teacher_assignment_constraints(assignment_rows: list[dict]) -> None:
    """Backward-compat shim — delegates to the structured applier.

    Kept as a free-function so callers that only want logging (no Course
    mutations) still work; today the call site uses the applier directly.
    """
    items = [(_safe_str(a.get("TEACHER_DCID")), _safe_str(a.get("LASTFIRST")),
              _safe_str(a.get("COURSENUMBER")), _safe_str(a.get("CONSTRAINTS")))
             for a in assignment_rows
             if _safe_str(a.get("CONSTRAINTS"))]
    if not items:
        return
    print(f"[INFO] teacher_assignments has {len(items)} CONSTRAINTS rows (free-text, not auto-enforced):")
    for dcid, name, course, ctext in items:
        print(f"    {course} / {name} ({dcid}): {ctext}")


def _read_requests(wb) -> list[dict]:
    # School renamed `requests` → `student_requests` on 2026-04-30 and added
    # two new columns (COURSENAME, TEARCHERASSISTANT). Fall back to the old
    # name for backward compat with older snapshots.
    sheet_name = "student_requests" if "student_requests" in wb.sheetnames else "requests"
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    out = []
    for raw in rows[1:]:
        if not raw or not raw[0]:
            continue
        d = dict(zip(headers, raw))
        if not d.get("COURSENUMBER") or not d.get("STUDENT_NUMBER"):
            continue
        out.append(d)
    return out


# ---------------------------------------------------------------------------
# Builder


def build_dataset_from_official_xlsx(
    xlsx_path: Path,
    grades: list[int] | None = None,
) -> Dataset:
    """Build a canonical Dataset from the official PS-exported xlsx.

    `grades` filters student requests to those grades. Default = full HS [9..12].

    Returns a Dataset where:
    - Course IDs are real PS COURSE_NUMBERs (e.g. 'OZ1313', 'E0901')
    - Teacher IDs are real PS DCIDs as strings (e.g. '377')
    - Room IDs are real PS room DCIDs as strings (e.g. '1', '304')
    - Student IDs are real PS STUDENT_NUMBERs (e.g. '28026')
    """
    if grades is None:
        grades = [9, 10, 11, 12]

    wb = load_workbook(xlsx_path, data_only=True)

    course_rows = _read_courses(wb)
    room_rows = _read_rooms(wb)
    teacher_rows = _read_teachers(wb)
    assignment_rows = _read_teacher_assignments(wb)
    request_rows = _read_requests(wb)

    # ------------------------------------------------------------------ rooms
    room_dcid_by_number: dict[str, str] = {}
    rooms: list[Room] = []
    for r in room_rows:
        dcid = _safe_str(r.get("DCID"))
        room_number = _safe_str(r.get("ROOMNUMBER"))
        department = _safe_str(r.get("DEPARTMENT"))
        capacity = _safe_int(r.get("MAXIMUM"), default=25) or 25
        room_dcid_by_number[room_number] = dcid
        rooms.append(Room(
            room_id=dcid,
            name=room_number,
            capacity=capacity,
            room_type=_infer_room_type(department, room_number),
            department=department or None,
        ))

    # --------------------------------------------------------------- courses
    courses: list[Course] = []
    course_by_number: dict[str, Course] = {}
    for c in course_rows:
        course_number = _safe_str(c["COURSE_NUMBER"])
        name = _safe_str(c.get("COURSE_NAME"))
        department = _safe_str(c.get("SCHED_DEPARTMENT"))
        max_size = _safe_int(c.get("MAXCLASSSIZE"), default=25) or 25
        # PE has MAXCLASSSIZE=130; AP Research has 26 explicitly.
        # Some courses have 0 → treat as default 25.
        if max_size <= 0:
            max_size = 25
        meetings = _safe_int(c.get("SCHED_FREQUENCY"), default=3) or 3
        is_advisory = course_number.upper().startswith("ADV")
        # Per client B6: only PE is curricularly "required". The rest is variable
        # but every student request must still be assigned (handled at request level).
        is_required_course = course_number.upper().startswith("E") and any(
            course_number.upper().endswith(suffix) for suffix in ("0901", "1001", "1101", "1201")
        )
        # MULTITERM 'S1' / 'S2' → semester course; everything else year-long
        multiterm = _safe_str(c.get("MULTITERM"))
        term = Term.SEMESTER if multiterm in ("S1", "S2") else Term.YEAR
        course = Course(
            course_id=course_number,
            name=name,
            department=department or "general",
            grade_eligibility=_infer_grades_from_name(name),
            is_required=is_required_course,
            meetings_per_week=meetings if not is_advisory else 1,
            max_size=max_size,
            is_advisory=is_advisory,
            qualified_teacher_ids=[],
            term=term,
        )
        courses.append(course)
        course_by_number[course_number] = course

    # ---------------------------------------------- course relationships (v4.4)
    # Prefer the in-file `course_relationships` tab; fall back to the legacy
    # sibling CSV if the tab is missing.
    relationships = _read_relationships_xlsx(wb)
    if not relationships:
        relationships = _read_relationships_csv(xlsx_path.parent / "course_relationships.csv")
    if relationships:
        _apply_course_relationships(relationships, courses, course_by_number)

    # ----------------------- teacher_assistants cleanup (v4.22, school 2026-04-30)
    # Strip TA placeholder entries from request_rows so the engine doesn't
    # fight with fake course assignments. School confirmed "teacher aide"
    # (course OZ1333) and "teacher assistant" (sheet teacher_assistants)
    # are equivalent.
    ta_records = _read_teacher_assistants_xlsx(wb)
    request_rows, ta_summary = _apply_teacher_assistants_cleanup(
        request_rows, ta_records, course_by_number
    )

    # -------------------------------------------------------------- teachers
    teachers: list[Teacher] = []
    teacher_by_dcid: dict[str, Teacher] = {}
    teacher_by_lastfirst: dict[str, Teacher] = {}
    for t in teacher_rows:
        dcid = _safe_str(t["DCID"])
        lastfirst = _safe_str(t.get("LASTFIRST"))
        department = _safe_str(t.get("SCHED_DEPARTMENT"))
        preferred_room_number = _safe_str(t.get("PREFERRED_ROOM"))
        home_room_id = room_dcid_by_number.get(preferred_room_number) if preferred_room_number else None
        teacher = Teacher(
            teacher_id=dcid,
            name=lastfirst,
            department=department or "general",
            qualified_course_ids=[],
            max_load=5,  # tightened below from observed assignments
            home_room_id=home_room_id,
        )
        teachers.append(teacher)
        teacher_by_dcid[dcid] = teacher
        teacher_by_lastfirst[lastfirst] = teacher

    # ----------------------- CONSTRAINTS column (auto-applies max-size overrides)
    # Must run BEFORE section creation so new sections inherit the bumped
    # Course.max_size. Free-text rules are still surfaced as [INFO].
    _apply_teacher_assignment_constraints(assignment_rows, courses, course_by_number)

    # ---------------------------------------------------- sections (from assignments)
    sections: list[Section] = []
    sections_per_teacher: dict[str, int] = defaultdict(int)
    section_counter_per_course: dict[str, int] = defaultdict(int)

    # v4.2: SCHEDULETERMCODE → Section.term_id (Columbus 2026-2027 mapping confirmed
    # by client 2026-04-28). Year-long courses keep term_id=None and inherit
    # SchoolConfig.term_id (3600) at export time.
    TERM_CODE_TO_ID = {"S1": "3601", "S2": "3602"}

    dropped_assignments: dict[str, list[tuple[str, str, int]]] = defaultdict(list)
    for a in assignment_rows:
        teacher_dcid = _safe_str(a["TEACHER_DCID"])
        course_number = _safe_str(a["COURSENUMBER"])
        n_sections = _safe_int(a.get("SECTIONS_PER_COURSE"), default=1)

        teacher = teacher_by_dcid.get(teacher_dcid)
        course = course_by_number.get(course_number)
        if teacher is None:
            dropped_assignments["teacher_not_in_teachers_sheet"].append(
                (teacher_dcid, course_number, n_sections)
            )
            continue
        if course is None:
            dropped_assignments["course_not_in_courses_sheet"].append(
                (teacher_dcid, course_number, n_sections)
            )
            continue

        # v4.3 — Term-paired sections (S1/S2 sharing slot).
        # SCHEDULETERMCODE "S1"/"S2" → semester section; map to term IDs
        # 3601 (S1) / 3602 (S2). Year-long sections keep term_id=None.
        term_code = _safe_str(a.get("SCHEDULETERMCODE"))
        section_term_id: str | None = TERM_CODE_TO_ID.get(term_code)
        if term_code and term_code not in ("26-27", "2026-2027") and section_term_id is None:
            continue

        # Wire up qualifications
        if course_number not in teacher.qualified_course_ids:
            teacher.qualified_course_ids.append(course_number)
        if teacher_dcid not in course.qualified_teacher_ids:
            course.qualified_teacher_ids.append(teacher_dcid)

        for _ in range(n_sections):
            section_counter_per_course[course_number] += 1
            idx = section_counter_per_course[course_number]
            sections.append(Section(
                section_id=f"{course_number}.{idx}",
                course_id=course_number,
                teacher_id=teacher_dcid,
                max_size=course.max_size,
                grade_level=course.grade_eligibility[0] if course.grade_eligibility else 12,
                term_id=section_term_id,
            ))
            if not course.is_advisory:
                sections_per_teacher[teacher_dcid] += 1

    # v4.2 — Simultaneous merge pass.
    # When a teacher is assigned to multiple courses sharing the same simul_group
    # (multi-level class: e.g. Spanish 9/10/11/12 FL, AP 2D + AP 3D Art), merge
    # those individual sections into ONE physical section with `linked_course_ids`
    # set to the additional courses. This fixes the apparent "overload" of
    # teachers like Clara Martínez (9 sections → 5) and Sofia Arcila.
    #
    # Strategy: for each (teacher, simul_group), pair sections by index — the
    # k-th section of each course in the group becomes one combined physical
    # section. If counts differ across courses (rare in practice), the extras
    # remain as standalone single-course sections.
    merged_groups: dict[tuple[str, str], dict[str, list[int]]] = {}
    for i, s in enumerate(sections):
        c = course_by_number.get(s.course_id)
        if c is None or c.simul_group is None:
            continue
        key = (s.teacher_id, c.simul_group)
        merged_groups.setdefault(key, {}).setdefault(s.course_id, []).append(i)

    to_remove: set[int] = set()
    for (tid, sg), by_course in merged_groups.items():
        if len(by_course) <= 1:
            continue  # only one course of the group present — nothing to merge
        # Pair by index: k-th section of each course becomes one combined section.
        ordered_courses = sorted(by_course.keys())  # determinism
        max_k = max(len(idxs) for idxs in by_course.values())
        for k in range(max_k):
            primary_idx: int | None = None
            for cid in ordered_courses:
                idxs = by_course[cid]
                if k >= len(idxs):
                    continue  # this course doesn't have a k-th section
                if primary_idx is None:
                    primary_idx = idxs[k]
                else:
                    # Append linked course to primary; mark this section for removal
                    primary = sections[primary_idx]
                    if cid != primary.course_id and cid not in primary.linked_course_ids:
                        primary.linked_course_ids.append(cid)
                    to_remove.add(idxs[k])
                    # Pigeonhole credit relief: this physical section is no longer
                    # "extra" load on the teacher.
                    if not course_by_number[cid].is_advisory:
                        sections_per_teacher[tid] -= 1
    if to_remove:
        sections = [s for i, s in enumerate(sections) if i not in to_remove]

    # Drop teachers with no academic sections (happens when SECTIONTYPE filter dropped all rows)
    # Note: keep all teachers — even those with only Advisory — since they appear in the data.

    # Apply per-teacher max_consec override for the pigeonhole-impossible cases
    for tid, n in sections_per_teacher.items():
        if n >= 7:
            t = teacher_by_dcid.get(tid)
            if t is not None:
                t.max_consecutive_classes = 5

    # Adjust max_load per observed
    for t in teachers:
        observed = sections_per_teacher.get(t.teacher_id, 0)
        if observed > t.max_load:
            t.max_load = observed

    # Drop courses that ended up with zero sections (no qualified teacher in this dataset)
    courses = [c for c in courses if c.is_advisory or any(s.course_id == c.course_id for s in sections)]
    course_by_number = {c.course_id: c for c in courses}

    # ----------------------------------------------------------------- students
    # v4.22 (school 2026-04-30): the new `required_courses` sheet defines
    # truly-required courses by grade. Only requests matching that map are
    # marked is_required=True; the rest are student-elected (still tracked
    # but the solver may drop them before dropping a real required).
    required_by_grade = _read_required_courses_xlsx(wb)
    if required_by_grade:
        n_total = sum(len(s) for s in required_by_grade.values())
        print(f"[INFO] required_courses: {n_total} truly-required (grade,course) pairs loaded")

    students_map: dict[str, Student] = {}
    n_marked_required = 0
    n_marked_elective = 0
    for r in request_rows:
        sid = _safe_str(r["STUDENT_NUMBER"])
        course_number = _safe_str(r["COURSENUMBER"])
        if not sid or not course_number:
            continue
        if course_number not in course_by_number:
            continue  # request for a course not in catalog — skip
        # Grade comes directly from the request row's STUDENT_GRADE_LEVEL_NEXT_YEAR.
        try:
            stu_grade = int(r.get("STUDENT_GRADE_LEVEL_NEXT_YEAR") or 12)
        except (TypeError, ValueError):
            stu_grade = 12
        if sid not in students_map:
            students_map[sid] = Student(
                student_id=sid,
                name=f"Student_{sid}",
                grade=stu_grade,  # use real grade from request row
                requested_courses=[],
            )
        # Mark is_required only if (grade, course) is in the official required map
        is_req = course_number in required_by_grade.get(stu_grade, set())
        if is_req:
            n_marked_required += 1
        else:
            n_marked_elective += 1
        students_map[sid].requested_courses.append(CourseRequest(
            student_id=sid,
            course_id=course_number,
            is_required=is_req,
            rank=1,
        ))
    if required_by_grade:
        print(f"[INFO] CourseRequest classification: {n_marked_required} HARD required, "
              f"{n_marked_elective} student-elected (electives)")

    # Add Advisory request to every student (Advisory is always-on for HS)
    advisory_id = next((c.course_id for c in courses if c.is_advisory), None)
    if advisory_id:
        for s in students_map.values():
            if not any(r.course_id == advisory_id for r in s.requested_courses):
                s.requested_courses.append(CourseRequest(
                    student_id=s.student_id,
                    course_id=advisory_id,
                    is_required=True,
                    rank=1,
                ))

    # Infer student grade from grade suffixes in their requested course names
    for s in students_map.values():
        grade_votes: dict[int, int] = defaultdict(int)
        for r in s.requested_courses:
            c = course_by_number.get(r.course_id)
            if not c:
                continue
            for g in c.grade_eligibility:
                if g in (9, 10, 11, 12):
                    grade_votes[g] += 1
        if grade_votes:
            # Pick the LOWEST grade with the most votes — usually their home grade.
            # (e.g. a grade-11 student takes some grade-10 courses; lowest is correct.)
            top = max(grade_votes.values())
            s.grade = min(g for g, v in grade_votes.items() if v == top)

    students = list(students_map.values())

    # Filter by requested grades if not full HS
    if set(grades) != {9, 10, 11, 12}:
        students = [s for s in students if s.grade in grades]

    # Advisory sections: prefer the PS-provided advisory rows from the assignment
    # sheet (canonical). Only synthesize new advisory sections if PS gave us none
    # — otherwise we duplicate the PS-derived ones and blow up HC1/HC4 budgets.
    if advisory_id:
        existing_adv = [s for s in sections if s.course_id == advisory_id]
        if not existing_adv:
            adv_size = course_by_number[advisory_id].max_size or 25
            if adv_size <= 0:
                adv_size = 25
            n_adv = max(6, -(-len(students) // adv_size))
            adv_teachers = [t for t in teachers if advisory_id in t.qualified_course_ids]
            if not adv_teachers:
                adv_teachers = teachers[:n_adv]
            for i in range(n_adv):
                tid = adv_teachers[i % len(adv_teachers)].teacher_id
                sections.append(Section(
                    section_id=f"{advisory_id}.{i+1}",
                    course_id=advisory_id,
                    teacher_id=tid,
                    max_size=adv_size,
                    grade_level=9,
                ))

    # Build SchoolConfig with the canonical PS values
    is_ms = all(g in (6, 7, 8) for g in grades)
    config = SchoolConfig(
        school="Columbus Middle School" if is_ms else "Columbus High School",
        school_id=12000 if is_ms else 13000,
        term_id="3600",
        grade=min(grades),
        year="2026-2027",
        bell=default_rotation(),
        hard=HardConstraints(),
        soft=SoftConstraintWeights(),
    )

    # ---- v4.4: parse the in-file tabs added by the school -----------------
    # co-planning groups (each row = teacher × course; group by COURSE ID).
    coplanning_groups = _read_coplanning_groups_xlsx(wb, teacher_by_lastfirst)
    if not coplanning_groups:
        # Fallback to the legacy sibling file
        coplanning_groups = _read_coplanning_groups(
            xlsx_path.parent / "rfi_1._STUDENTS_PER_COURSE_2026-2027.xlsx",
            teacher_by_lastfirst,
        )

    # conselours_recommendations → BehaviorMatrix (separations + groupings)
    valid_student_ids = set(students_map.keys())
    behavior_matrix, name_to_id = _read_behavior_matrix_xlsx(wb, valid_student_ids)

    # teacher_avoid → Student.restricted_teacher_ids (uses name_to_id from above)
    avoid_matched = _read_teacher_avoid_xlsx(
        wb, students_map, teacher_by_lastfirst, name_to_id
    )

    # CONSTRAINTS column was already applied above (max-size overrides) and
    # the free-text rules logged at that time. Nothing else to do here.

    print(
        f"[INFO] ingested {len(coplanning_groups)} co-planning groups, "
        f"{len(behavior_matrix.separations)} separations, "
        f"{len(behavior_matrix.groupings)} groupings, "
        f"{avoid_matched} teacher_avoid restrictions"
    )

    # ---- Data-quality summary (visible in build output) -------------------
    # 1) Assignments dropped because of missing teacher / course rows.
    if dropped_assignments:
        print("[WARN] dropped teacher_assignments rows during ingest:")
        for reason, items in dropped_assignments.items():
            n_rows = len(items)
            n_secs = sum(s for _, _, s in items)
            print(f"  {reason}: {n_rows} rows ({n_secs} sections)")
            for tdcid, cnum, nsec in items[:10]:
                print(f"    teacher_dcid={tdcid!r} course={cnum!r} sections={nsec}")
            if n_rows > 10:
                print(f"    … +{n_rows - 10} more")

    # 2) Cross-check planned (SECTIONSTOOFFER) vs actual sections per course,
    #    counting linked_course_ids so multi-level merges don't trigger false
    #    alarms. We treat a mismatch as a WARN, not an error — the school may
    #    legitimately under-supply teachers (capacity-bound courses).
    planned_by_course: dict[str, int] = {}
    for c in course_rows:
        cnum = _safe_str(c.get("COURSE_NUMBER"))
        if cnum:
            planned_by_course[cnum] = _safe_int(c.get("SECTIONSTOOFFER"), default=0)
    actual_by_course: dict[str, int] = defaultdict(int)
    for s in sections:
        actual_by_course[s.course_id] += 1
        for linked in s.linked_course_ids:
            actual_by_course[linked] += 1
    mismatches = []
    for cnum, planned in planned_by_course.items():
        actual = actual_by_course.get(cnum, 0)
        if planned != actual and (planned > 0 or actual > 0):
            mismatches.append((cnum, planned, actual))
    if mismatches:
        print("[WARN] courses where planned (SECTIONSTOOFFER) != generated sections:")
        for cnum, planned, actual in sorted(mismatches, key=lambda x: -abs(x[1] - x[2]))[:20]:
            cn = course_by_number.get(cnum)
            name = cn.name if cn else "??"
            print(f"  {cnum:10s} {name:35s} planned={planned}  generated={actual}  diff={planned-actual:+d}")
        if len(mismatches) > 20:
            print(f"  … +{len(mismatches) - 20} more")

    # 3) Demand vs capacity per course — flag where requests > sum(max_size).
    demand_by_course: dict[str, int] = defaultdict(int)
    for r in request_rows:
        cnum = _safe_str(r.get("COURSENUMBER"))
        if cnum:
            demand_by_course[cnum] += 1
    capacity_by_course: dict[str, int] = defaultdict(int)
    for s in sections:
        capacity_by_course[s.course_id] += s.max_size
        for linked in s.linked_course_ids:
            capacity_by_course[linked] += s.max_size
    deficits = []
    for cnum, demand in demand_by_course.items():
        cap = capacity_by_course.get(cnum, 0)
        if demand > cap:
            deficits.append((cnum, demand, cap, demand - cap))
    if deficits:
        print("[WARN] courses with demand > capacity (capacity-bound, will produce unmet):")
        for cnum, demand, cap, deficit in sorted(deficits, key=lambda x: -x[3])[:20]:
            cn = course_by_number.get(cnum)
            name = cn.name if cn else "??"
            print(f"  {cnum:10s} {name:35s} demand={demand}  cap={cap}  short={deficit}")
        if len(deficits) > 20:
            print(f"  … +{len(deficits) - 20} more")

    return Dataset(
        config=config,
        courses=courses,
        teachers=teachers,
        rooms=rooms,
        sections=sections,
        students=students,
        behavior=behavior_matrix,
        coplanning_groups=coplanning_groups,
    )
